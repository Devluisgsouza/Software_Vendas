import openpyxl as op
from copy import copy

exit = False

def criar_conta():
    arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
    banco_contas = arquivo_bancodd["contas_de_usuario"]
    ultima_linha = banco_contas.max_row
    for linha in range(2, ultima_linha +1):
        ultima_linha = linha
    id_origem = banco_contas.cell(row=2, column=1)
    email_origem = banco_contas.cell(row=2, column=2)
    senha_origem = banco_contas.cell(row=2, column=3)
    print(f"{"\n"}{"\033[1;36m CREATE YOUR ACCOUNT ".center(52)}{"\n"}")
    while True:
        if exit:
            break
        login = input("\033[1;32mEmail: ").strip().lower()
        password = input("\033[1;32mPassword: ").strip()
        for linha in range(2,ultima_linha + 1):
            emails = banco_contas.cell(row=linha, column=2).value
            id = banco_contas.cell(row=linha, column=1).value
            id_conta = int(id)  + 1
            if login == emails:
                print(f"{"\n"}{"\033[1;31m THIS ACCOUNT ALREADY EXISTS IN THE SYSTEM \033[m".center(52, "=")}{"\n"}")
                break
        else:
            print(f"{"\n"}{"\033[1;32m ACCOUNT CREATED SUCCESSFULY! \033[m".center(54)}")
            banco_contas.append([id_conta,login,password])
            email_destino = banco_contas.cell(row=ultima_linha+1, column=2)
            senha_destino = banco_contas.cell(row=ultima_linha+1, column=3)
            id_destino = banco_contas.cell(row=ultima_linha+1, column=1)
            email_destino._style = copy(email_origem._style)
            senha_destino._style = copy(senha_origem._style)
            id_destino._style = copy(id_origem._style)
            arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
            return exit
 

def fazer_login():
    arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
    banco_contas = arquivo_bancodd["contas_de_usuario"]
    ultima_linha = banco_contas.max_row
    print(f"{"\n"}{"\033[1;36m ENTER WITH YOUR ACCOUNT: ".center(52)}{"\n"}")
    login = input("\033[1;32mEmail: ").strip()
    password = input("Password: ").strip()
    for linha in range(2,ultima_linha + 1):
        emails = banco_contas.cell(row=linha, column=2).value
        if login == emails:
            numero_linha_logins = banco_contas.cell(row=linha, column=2).row
            senhas = banco_contas.cell(row=numero_linha_logins, column=3).value
            if password == senhas:
                print(f"{"\n"}{"\033[1;32m SUCCESSFULLY LOGIN \033[m".center(51)}")
                return
    print(f"{"\n"}{"\033[1;31m WRONG INFORMATIONS! \033[m".center(55)}")
    print(f"\033[1;31mTHIS ACCOUNT DOESN'T EXIST IN THE SYSTEM".center(52))
    while True:
        print(f"{"\n"}{"\033[1;36m WOULD YOU LIKE TO CREATE A NEW ACCOUNT? ".center(53)}{"\n"}")
        resp = input(f"{" [1]YES   OR   [2]NO ".center(44)}{"\n"}{"\n"}").replace(" ","").lower()
        if resp == "1":
            criar_conta()
            break
        elif resp == "2":
            fazer_login()
            break
        else:
            print(f"{"\n"}{"\033[1;31mI CAN'T UNDERSTAND WHAT YOU WANT\033[m".center(53)}")
            print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(53)}{"\n"}")
            

import openpyxl as op

def exlcuir_venda():
    exit = False
    while True:
        if exit:
            break
        arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
        banco_vendas = arquivo_bancodd["banco_de_vendas"]
        ultima_linha = banco_vendas.max_row
        idstr = input(f"\033[1;36m{"\n"}{"ENTER THE “ID” OF THE SALE YOU WANT TO DELETE ".center(48)}{"\n"}{"\n"}").lower().strip()
        try:
            id = int(idstr)
        except:
            print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID ID \033[m".center(51)}")
            print(f"{"\033[1;31m USE ONLY NUMBERS TO ENTER THE ID \033[m".center(52)}")
            continue
        for linha in range(2,ultima_linha + 1):
            id_venda = banco_vendas.cell(row=linha, column=1).value
            if id == id_venda:
                print(f"{"\n"}{"\n"}{"\033[1;31mARE YOU SURE?".center(52)}{"\n"}")
                decide = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip().lower()
                if decide == "1":
                    print(f"{"\n\033[1;32m"}{"SALE SUCCESSFULLY DELETED".center(46)}{"\n"}")
                    banco_vendas.delete_rows(linha,1)
                elif decide == "2":
                    break
                else:
                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                    break                    
                arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                
                while True:
                    print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO DELETE MORE SALES?".center(53)}")
                    resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                    if resp == "1":
                        break
                    elif resp == "2":
                        return exit
                    else:
                        print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                        print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")

                

def alterar_venda():
    arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
    banco_vendas = arquivo_bancodd["banco_de_vendas"]
    ultima_linha = banco_vendas.max_row
    
    exit = False

    while True:
        if exit:
            break
        idstr = input(f"\033[1;36m{"\n"}{"ENTER THE “ID” OF THE SALE YOU WANT TO CHANGE ".center(48)}{"\n"}{"\n"}").lower().strip()
        id = int(idstr)
        for linha in range(2,ultima_linha + 1):
            id_venda = banco_vendas.cell(row=linha, column=1).value
            if id == id_venda:
                while True:
                    print(f"{"\n"}{"\033[1;36mDO YOU WANT TO CHANGE:  ".center(53)}{"\n"}")
                    print("[1] PRODUCT NAME ".center(39))
                    print("[2] PRODUCT PRICE ".center(39))
                    print("[3] DATE OF SALE ".center(39))
                    print("[4] RETURN TO MENU ".center(41))
                    resp = input("\n").strip()
                    match resp:
                        case "1":
                            while True:
                                new_name = input(f"{"\033[1;36mENTER WITH THE NEW NAME: \033[1;32m".center(59)}{"\n"*2}").strip().lower()
                                if new_name == "":
                                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID NAME \033[m".center(54)}{"\n"}")
                                else:
                                    print(f"{"\n"}{"\033[1;32m"}{"NAME SUCCESSFULLY CHANGED".center(45)}{"\n"}")
                                    banco_vendas.cell(row=linha, column=2).value = new_name
                                    arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                                    break
                            while True:
                                print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO CHANGE MORE SALES?".center(53)}")
                                resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                                if resp == "1":
                                    break
                                elif resp == "2":
                                    return exit
                                else:
                                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                            break
                        case "2":
                            while True:
                                pricestr = input(f"{"\033[1;36mENTER WITH THE NEW PRICE: \033[1;32m".center(58)}{"\n"*2}{"R$"}").strip().lower().replace(",",".")
                                try:
                                    new_price = float(pricestr)
                                    banco_vendas.cell(row=linha, column=3).value = new_price
                                    print(f"{"\n"}{"\033[1;32m"}{"PRICE SUCCESSFULLY CHANGED".center(45)}{"\n"}")
                                    arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                                    break
                                except:
                                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID PRICE \033[m".center(52)}")
                                    print(f"{"\033[1;31m USE ONLY THE “,” OR “.” TO ADD CENTS \033[m".center(52)}")
                                    print(f"{"\n"}{"\033[1;31m FOR EXAMPLE:\033[m".center(52)}")
                                    print(f"{"\033[1;32mR$0000,00 or R$0000.00\033[m".center(51)}{"\n"}")
                            while True:
                                print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO CHANGE MORE SALES?".center(53)}")
                                resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                                if resp == "1":
                                    break
                                elif resp == "2":
                                    return exit
                                else:
                                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                            break
                        case "3":
                            while True:
                                print(f"{"\n"}{"\033[1;36mENTER WITH THE NEW DATE: \033[1;32m".center(59)}")
                                print(f"{"\n"}{"FOR EXAMPLE: DD/MM/YYYY".center(44)}{"\n"}")
                                date = input().replace("-","/").replace(".","/").replace(" ","/").strip()
                                if date == "" or " ":
                                    print(f"{"\033[1;31m ERROR! ENTER A VALID DATE \033[m".center(53)}{"\n"}")
                                else:
                                    print(f"{"\n"}{"\033[1;32m"}{"DATE SUCCESSFULLY CHANGED".center(45)}{"\n"}")
                                    banco_vendas.cell(row=linha, column=4).value = date
                                    arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                                    break               
                            while True:
                                print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO CHANGE MORE SALES?".center(53)}")
                                resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                                if resp == "1":
                                    break
                                elif resp == "2":
                                    return exit
                                else:
                                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                            break
                        case "4":
                            return exit   
                        case _:
                            print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT! ".center(50)}")
                            print(f"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2,3,4]\033[m]".center(47))                                                            
            
        

import openpyxl as op


exit = False

def vendashist():
    arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
    banco_vendas = arquivo_bancodd["banco_de_vendas"]
    ultima_linha = banco_vendas.max_row
    for linha in range(2,ultima_linha + 1):
        produtos = banco_vendas.cell(row=linha, column=2).value
        preços = banco_vendas.cell(row=linha, column=3).value
        datas = banco_vendas.cell(row=linha, column=4).value
    
    while True:
        if exit:
            break
        print(f"{"\n"}{"\033[1;34mSALES HISTORY".center(42)}\n")
        print(f"{"DO YOU WANT TO SEARCH BY:  ".center(49)}{"\n"}")
        print("[1] PRODUCT NAME ".center(39))
        print("[2] PRODUCT PRICE ".center(39))
        print("[3] DATE OF SALE ".center(39))
        print("[4] RETURN TO MENU ".center(41))
        resp = input("\n")
        while True:
            if resp == "1":
                resp2 = input(f"{"\n"}{"\033[1;36mENTER THE PRODUCT NAME: ".center(50)}{"\n"}{"\n"}\033[1;32m").strip().lower()
                if resp2 != "":
                    print(f"{"\n"}{"\033[1;36mPRODUCT HISTORY WITH NAME: ".center(47)}\033[1;32m{resp2}{"\n"}")
                    for linha in range(2,ultima_linha + 1):
                        produtos = banco_vendas.cell(row=linha, column=2).value
                        preços = banco_vendas.cell(row=linha, column=3).value
                        datas = banco_vendas.cell(row=linha, column=4).value
                        if resp2 == produtos:
                            print(f"{"PRODUCT: "}{produtos}{"  /  PRICE: R$"}{preços:.2f}{"  /  SALE DATE: "}{datas}")  
                    break                  
                else:
                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID NAME \033[m".center(51)}{"\n"}")

            elif resp == "2":
                resp3 = input(f"{"\n"}{"\033[1;36mENTER THE PRODUCT PRICE: ".center(50)}{"\n"}{"\n"}\033[1;32m").replace(",", ".").strip()
                try:
                    resp3float = float(resp3)
                    print(f"{"\n"}{"\033[1;36mPRODUCT HISTORY WITH PRICE: ".center(43)}\033[1;32m{"R$"}{resp3float:.2f}{"\n"}")    
                    for linha in range(2,ultima_linha + 1):
                        produtos = banco_vendas.cell(row=linha, column=2).value
                        preços = banco_vendas.cell(row=linha, column=3).value
                        datas = banco_vendas.cell(row=linha, column=4).value
                        if resp3float == preços:
                            print(f"{"PRODUCT: "}{produtos}{"  /  PRICE: R$"}{preços:.2f}{"  /  SALE DATE: "}{datas}")  
                    break    
                except ValueError:
                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID PRICE \033[m".center(52)}{"\n"}")   
            elif resp == "3":
                print(f"{"\n"}{"\033[1;36mENTER THE PRODUCT DATE: ".center(50)}{"\n"}\033[1;32m")
                resp4 = input("\033[1;36mFOR EXAMPLE: DD/MM/YYYY (DAY/MONTH/YEAR)\n\n".center(53)).replace("-", "/").replace(" ","/").replace(".","/").strip()
                print(f"{"\n"}{"\033[1;36mPRODUCT HISTORY WITH DATE: ".center(47)}\033[1;32m{resp4}{"\n"}")   
                print(f"{"\n"}{"\033[1;36mNAME , PRICE , DATE\033[m\n\033[1;32m"}")  
                for linha in range(2,ultima_linha + 1):
                    produtos = banco_vendas.cell(row=linha, column=2).value
                    preços = banco_vendas.cell(row=linha, column=3).value
                    datas = banco_vendas.cell(row=linha, column=4).value
                    if resp4 == datas:
                        print(f"{"PRODUCT: "}{produtos}{"  /  PRICE: R$"}{preços:.2f}{"  /  SALE DATE: "}{datas}")  
                break  
            elif resp == "4":
                return exit
            else:
                print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT! ".center(49)}")
                print(f"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2,3,4]".center(47))
                break
                


def menusystem():
    print(f"{"\n"}{"\033[1;36m SYSTEM MENU" .center(40)}{"\n"}")
    print(" [1] ACCESS THE SALES SYSTEM ".center(49))
    print(" [2] CHANGE ACCOUNT ".center(39))
    print(" [3] ACCESS THE SALES HISTORY ".center(50))
    print(" [4] DELETE SALES ".center(38))
    print(" [5] CHANGE SALES ".center(38))
    print(f"{" [6] CLOSE SYSTEM ".center(37)}{"\n"}")


def menusystem2():
    print(f"{"\n"}\033[1;36m{" WHAT YOU WANT TO DO? ".center(41)}")
    print(f"{"\n"}{"[1] NEW SALE".center(33)}")
    print(" [2] CHANGE ACCOUNT ".center(40))
    print(" [3] ACCESS THE SALES HISTORY ".center(50))
    print(" [4] DELETE SALES ".center(38))
    print(" [5] CHANGE SALES ".center(38))
    print(f"{" [6] CLOSE SYSTEM ".center(37)}{"\n"}")


from datetime import datetime
import openpyxl as op
from copy import copy

tempo = datetime.now().date()
data = tempo.strftime("%d/%m/%Y")
exit = False

def vender_prod():
    print(f"{"\n"}{"\033[1;36m WELCOME TO SALES SYSTEM! ".center(50)}{"\n"}")
    print(f"{"\033[1;36m PUT ALL THE PRODUCTS YOU'VE SOLD ".center(50)}{"\n"}")
    print(f"{"\n"}{"\033[1;34m PRODUCTS OVER R$100.00 HAVE A 10% DISCOUNT! ".center(50)}{"\n"}")
    print(f"{"IF YOU WANT TO RETURN TO MENU, TYPE 'EXIT'".center(44)}{"\n"}")
    
    total = 0
    totprod = 0
    totdesc = 0
    sacola = []

    while True:
        if exit:
                break
        arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
        banco_vendas = arquivo_bancodd["banco_de_vendas"]
        ultima_linha_base = banco_vendas.max_row
        for linha in range(2, ultima_linha_base +1):
            ultima_linha = linha

        id = banco_vendas.cell(row=linha, column=1).value
        id_venda = int(id) + 1
        id_origem = banco_vendas.cell(row=3, column=1)
        nome_origem = banco_vendas.cell(row=3, column=2)
        preço_origem = banco_vendas.cell(row=3, column=3)
        data_origem = banco_vendas.cell(row=3, column=4)
        
        
        while True:
            name = input(f"\033[1;36m{"\n"}{"Product's name: "}").lower().strip()
            if name == "":
                print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID NAME \033[m".center(51)}{"\n"}")
            elif name == "exit".strip().lower():
                return exit
            else:
                break
        while True:
            pricestr = input("\033[1;36mProduct's price: R$").replace(",",".").strip().lower()
            if pricestr == "exit".strip().lower():
                return exit
            try:
                price = float(pricestr)
                break
            except:
                print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID PRICE \033[m".center(52)}")
                print(f"{"\033[1;31m USE ONLY THE “,” OR “.” TO ADD CENTS \033[m".center(52)}")
                print(f"{"\n"}{"\033[1;31m FOR EXAMPLE:\033[m".center(52)}")
                print(f"{"\033[1;32mR$0000,00 or R$0000.00\033[m".center(51)}{"\n"}")            
        if price < 100:
            total += price
            totprod += 1
            sacola.append((name, price))
            banco_vendas.append([id_venda,name,price,data])
            id_destino = banco_vendas.cell(row=ultima_linha+1, column=1)
            preço_destino = banco_vendas.cell(row=ultima_linha+1, column=3)
            data_destino = banco_vendas.cell(row=ultima_linha+1, column=4)
            nome_destino = banco_vendas.cell(row=ultima_linha+1, column=2)
            id_destino._style = copy(id_origem._style)
            preço_destino._style = copy(preço_origem._style)
            data_destino._style = copy(data_origem._style)
            nome_destino._style = copy(nome_origem._style)
            arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
        else:
            price = price - (price * 0.1)
            totdesc += 1
            total += price
            totprod += 1
            sacola.append((name, price))
            banco_vendas.append([id_venda,name,price,data])
            id_destino = banco_vendas.cell(row=ultima_linha+1, column=1)
            preço_destino = banco_vendas.cell(row=ultima_linha+1, column=3)
            data_destino = banco_vendas.cell(row=ultima_linha+1, column=4)
            nome_destino = banco_vendas.cell(row=ultima_linha+1, column=2)
            id_destino._style = copy(id_origem._style)
            preço_destino._style = copy(preço_origem._style)
            data_destino._style = copy(data_origem._style)
            nome_destino._style = copy(nome_origem._style)
            arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
        while True:
            print(f"{"\n"}{"\033[1;36m ADD MORE PRODUCTS? ".center(48)}{"\n"}")
            decide = input(f"{" [1]YES   OR   [2]NO ".center(42)}{"\n"}{"\n"}").strip().lower()
            if decide == "2":
                print(f"{"\n"}{"\033[1;33m SUMMARY OF YOUR SALE ".center(48)}")
                print(
                    f"{"\n"}{"\033[1;33mTotal products: ".ljust(35)}\033[1;32m{totprod:.1f}\033[m")
                print(
                    f"{"\033[1;33mTotal products on sale: ".ljust(35)}\033[1;32m{totdesc:.1f}\033[m")
                print(
                    f"{"\033[1;33mTotal to pay: ".ljust(35)}\033[1;32mR${total:.2f}\033[m")
                print(f"{"\n"}{"\033[1;33m ALL PRODUCTS SOLD ".center(47)}{"\n"}")
                for name, price in sacola:
                    print(f"\033[1;33m{name}{"    -    "}\033[1;32mR${price:.2f}")
                print(f"{"\n"}{"\033[m"}")
                arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                break
            elif decide == "1":
                arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                break
            else:
                print(
                    f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT ".center(50)}")
                print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(52)}{"\n"}")
        if decide != "1" and "2":
            break
    