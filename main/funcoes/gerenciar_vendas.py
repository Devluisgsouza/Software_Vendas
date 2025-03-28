import openpyxl as op

def exlcuir_venda():
    exit = False
    while True:
        if exit:
            break
        arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
        banco_vendas = arquivo_bancodd["banco_de_vendas"]
        ultima_linha = banco_vendas.max_row
        idstr = input(f"\033[1;36m{"\n"}{"ENTER THE “ID” OF THE SALE YOU WANT TO DELETE ".center(48)}{"\n"}{"\n"}").lower().strip()
        try:
            id = int(idstr)
        except:
            print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID ID \033[m".center(51)}")
            print(f"{"\033[1;31m USE ONLY NUMBERS TO ENTER THE ID \033[m".center(52)}")
            continue
        for linha in range(2,ultima_linha + 1):
            id_venda = banco_vendas.cell(row=linha, column=1).value
            if id == id_venda:
                print(f"{"\n"}{"\n"}{"\033[1;31mARE YOU SURE?".center(52)}{"\n"}")
                decide = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip().lower()
                if decide == "1":
                    print(f"{"\n\033[1;32m"}{"SALE SUCCESSFULLY DELETED".center(46)}{"\n"}")
                    banco_vendas.delete_rows(linha,1)
                elif decide == "2":
                    break
                else:
                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                    break                    
                arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                
                while True:
                    print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO DELETE MORE SALES?".center(53)}")
                    resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                    if resp == "1":
                        break
                    elif resp == "2":
                        return exit
                    else:
                        print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                        print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")

                

def alterar_venda():
    arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
    banco_vendas = arquivo_bancodd["banco_de_vendas"]
    ultima_linha = banco_vendas.max_row
    
    exit = False

    while True:
        if exit:
            break
        idstr = input(f"\033[1;36m{"\n"}{"ENTER THE “ID” OF THE SALE YOU WANT TO CHANGE ".center(48)}{"\n"}{"\n"}").lower().strip()
        id = int(idstr)
        for linha in range(2,ultima_linha + 1):
            id_venda = banco_vendas.cell(row=linha, column=1).value
            if id == id_venda:
                while True:
                    print(f"{"\n"}{"\033[1;36mDO YOU WANT TO CHANGE:  ".center(53)}{"\n"}")
                    print("[1] PRODUCT NAME ".center(39))
                    print("[2] PRODUCT PRICE ".center(39))
                    print("[3] DATE OF SALE ".center(39))
                    print("[4] RETURN TO MENU ".center(41))
                    resp = input("\n").strip()
                    match resp:
                        case "1":
                            while True:
                                new_name = input(f"{"\033[1;36mENTER WITH THE NEW NAME: \033[1;32m".center(59)}{"\n"*2}").strip().lower()
                                if new_name == "":
                                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID NAME \033[m".center(54)}{"\n"}")
                                else:
                                    print(f"{"\n"}{"\033[1;32m"}{"NAME SUCCESSFULLY CHANGED".center(45)}{"\n"}")
                                    banco_vendas.cell(row=linha, column=2).value = new_name
                                    arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                                    break
                            while True:
                                print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO CHANGE MORE SALES?".center(53)}")
                                resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                                if resp == "1":
                                    break
                                elif resp == "2":
                                    return exit
                                else:
                                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                            break
                        case "2":
                            while True:
                                pricestr = input(f"{"\033[1;36mENTER WITH THE NEW PRICE: \033[1;32m".center(58)}{"\n"*2}{"R$"}").strip().lower().replace(",",".")
                                try:
                                    new_price = float(pricestr)
                                    banco_vendas.cell(row=linha, column=3).value = new_price
                                    print(f"{"\n"}{"\033[1;32m"}{"PRICE SUCCESSFULLY CHANGED".center(45)}{"\n"}")
                                    arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                                    break
                                except:
                                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID PRICE \033[m".center(52)}")
                                    print(f"{"\033[1;31m USE ONLY THE “,” OR “.” TO ADD CENTS \033[m".center(52)}")
                                    print(f"{"\n"}{"\033[1;31m FOR EXAMPLE:\033[m".center(52)}")
                                    print(f"{"\033[1;32mR$0000,00 or R$0000.00\033[m".center(51)}{"\n"}")
                            while True:
                                print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO CHANGE MORE SALES?".center(53)}")
                                resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                                if resp == "1":
                                    break
                                elif resp == "2":
                                    return exit
                                else:
                                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                            break
                        case "3":
                            while True:
                                print(f"{"\n"}{"\033[1;36mENTER WITH THE NEW DATE: \033[1;32m".center(59)}")
                                print(f"{"\n"}{"FOR EXAMPLE: DD/MM/YYYY".center(44)}{"\n"}")
                                date = input().replace("-","/").replace(".","/").replace(" ","/").strip()
                                if date == "" or " ":
                                    print(f"{"\033[1;31m ERROR! ENTER A VALID DATE \033[m".center(53)}{"\n"}")
                                else:
                                    print(f"{"\n"}{"\033[1;32m"}{"DATE SUCCESSFULLY CHANGED".center(45)}{"\n"}")
                                    banco_vendas.cell(row=linha, column=4).value = date
                                    arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                                    break               
                            while True:
                                print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO CHANGE MORE SALES?".center(53)}")
                                resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                                if resp == "1":
                                    break
                                elif resp == "2":
                                    return exit
                                else:
                                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                            break
                        case "4":
                            return exit   
                        case _:
                            print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT! ".center(50)}")
                            print(f"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2,3,4]\033[m]".center(47))                                                            
            
        

# alterar_venda()
# exlcuir_venda()
