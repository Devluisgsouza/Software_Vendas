import openpyxl as op

def excluir_conta():
    exit = False
    while True:
        if exit:
            break
        arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
        banco_contas = arquivo_bancodd["contas_de_usuario"]
        ultima_linha = banco_contas.max_row
        idstr = input(f"\033[1;36m{"\n"}{"ENTER THE “ID” OF THE ACCOUNT YOU WANT TO DELETE ".center(48)}{"\n"}{"\n"}").lower().strip()
        try:
            id = int(idstr)
        except:
            print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID ID \033[m".center(51)}")
            print(f"{"\033[1;31m USE ONLY NUMBERS TO ENTER THE ID \033[m".center(52)}")
            continue
        for linha in range(2,ultima_linha + 1):
            id_venda = banco_contas.cell(row=linha, column=1).value
            if id == id_venda:
                print(f"{"\n"}{"\n"}{"\033[1;31mARE YOU SURE?".center(52)}{"\n"}")
                decide = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip().lower()
                if decide == "1":
                    print(f"{"\n\033[1;32m"}{"ACCOUNT SUCCESSFULLY DELETED".center(46)}{"\n"}")
                    banco_contas.delete_rows(linha,1)
                elif decide == "2":
                    break
                else:
                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                    break                    
                arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                
                while True:
                    print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO DELETE MORE ACCOUNTS?".center(53)}")
                    resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                    if resp == "1":
                        break
                    elif resp == "2":
                        return exit
                    else:
                        print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                        print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")



def alterar_conta():  
    exit = False
    while True:
        if exit:
            break
        arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
        banco_contas = arquivo_bancodd["contas_de_usuario"]
        ultima_linha = banco_contas.max_row        
        idstr = input(f"\033[1;36m{"\n"}{"ENTER THE “ID” OF THE SALE YOU WANT TO CHANGE ".center(48)}{"\n"}{"\n"}").lower().strip()
        id = int(idstr)
        for linha in range(2,ultima_linha + 1):
            id_venda = banco_contas.cell(row=linha, column=1).value
            if id == id_venda:
                while True:
                    print(f"{"\n"}{"\033[1;36mDO YOU WANT TO CHANGE:  ".center(53)}{"\n"}")
                    print("[1] ACCOUNT EMAIL ".center(39))
                    print("[2] ACCOUNT PASSWORD ".center(43))
                    print("[3] RETURN TO MENU ".center(41))
                    resp = input("\n").strip()
                    match resp:
                        case "1":
                            while True:
                                new_email = input(f"{"\033[1;36mENTER WITH THE NEW EMAIL: \033[1;32m".center(59)}{"\n"*2}").strip().lower()
                                if new_email == "":
                                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID EMAIL \033[m".center(54)}{"\n"}")
                                else:
                                    print(f"{"\n"}{"\033[1;32m"}{"EMAIL SUCCESSFULLY CHANGED".center(45)}{"\n"}")
                                    banco_contas.cell(row=linha, column=2).value = new_email
                                    arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                                    break
                            while True:
                                print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO CHANGE MORE ACCOUNTS?".center(53)}")
                                resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                                if resp == "1":
                                    break
                                elif resp == "2":
                                    return exit
                                else:
                                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                            break
                        case "2":
                            while True:
                                new_password = input(f"{"\033[1;36mENTER WITH THE NEW PASSWORD: \033[1;32m".center(59)}{"\n"*2}").strip()
                                if new_password == "":
                                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID PASSWORD \033[m".center(54)}{"\n"}")
                                else:
                                    print(f"{"\n"}{"\033[1;32m"}{"PASSWORD SUCCESSFULLY CHANGED".center(45)}{"\n"}")
                                    banco_contas.cell(row=linha, column=3).value = new_password
                                    arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                                    break
                            while True:
                                print(f"{"\n"}{"\033[1;36mWOULD YOU LIKE TO CHANGE MORE ACCOUNTS?".center(53)}")
                                resp = input(f"{'[1]YES   OR   [2]NO'.center(45)}{"\n"}").strip()
                                if resp == "1":
                                    break
                                elif resp == "2":
                                    return exit
                                else:
                                    print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT \033[m".center(56)}")
                                    print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(54)}{"\n"}")
                            break
                        case "3":
                            return exit   
                        case _:
                            print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT! ".center(50)}")
                            print(f"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2,3]\033[m]".center(47))      





# excluir_conta()
# alterar_conta()
