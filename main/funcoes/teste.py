import openpyxl as op



arquivo_bandodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
banco = arquivo_bandodd["banco_de_dados"]
ultima_linha = banco.max_row


for linha in range(2, ultima_linha +1):
    produto = banco.cell(row=linha, column=1).value
    if produto == "a":
        print(produto)

























arquivo_bandodd.save("main\\banco_de_dados\\banco.xlsx")