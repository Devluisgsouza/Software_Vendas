from datetime import datetime
import openpyxl as op
from copy import copy

tempo = datetime.now().date()
data = tempo.strftime("%d/%m/%Y")
exit = False

def vender_prod():
    print(f"{"\n"}{"\033[1;36m WELCOME TO SALES SYSTEM! ".center(50)}{"\n"}")
    print(f"{"\033[1;36m PUT ALL THE PRODUCTS YOU'VE SOLD ".center(50)}{"\n"}")
    print(f"{"\n"}{"\033[1;34m PRODUCTS OVER R$100.00 HAVE A 10% DISCOUNT! ".center(50)}{"\n"}")
    print(f"{"IF YOU WANT TO RETURN TO MENU, TYPE 'EXIT'".center(44)}{"\n"}")
    
    total = 0
    totprod = 0
    totdesc = 0
    sacola = []

    while True:
        if exit:
                break
        
        arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
        banco = arquivo_bancodd["banco_de_dados"]
        ultima_linha_base = banco.max_row
        for linha in range(2, ultima_linha_base +1):
            ultima_linha = linha
        nome_origem = banco.cell(row=3, column=1)
        preço_origem = banco.cell(row=3, column=2)
        data_origem = banco.cell(row=3, column=3)
        
        while True:
            name = input(f"\033[1;36m{"\n"}{"Product's name: "}").lower().strip()
            if name == "":
                print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID NAME \033[m".center(51)}{"\n"}")
            elif name == "exit".strip().lower():
                return exit
            else:
                break
        while True:
            pricestr = input("\033[1;36mProduct's price: R$").replace(",",".").strip().lower()
            if pricestr == "exit".strip().lower():
                return exit
            try:
                price = float(pricestr)
                break
            except:
                print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID PRICE \033[m".center(52)}")
                print(f"{"\033[1;31m USE ONLY THE “,” OR “.” TO ADD CENTS \033[m".center(52)}")
                print(f"{"\n"}{"\033[1;31m FOR EXAMPLE:\033[m".center(52)}")
                print(f"{"\033[1;32mR$0000,00 or R$0000.00\033[m".center(51)}{"\n"}")
                     
        if price < 100:
            total += price
            totprod += 1
            sacola.append((name, price))
            banco.append([name,price,data])
            preço_destino = banco.cell(row=ultima_linha+1, column=2)
            data_destino = banco.cell(row=ultima_linha+1, column=3)
            nome_destino = banco.cell(row=ultima_linha+1, column=1)
            preço_destino._style = copy(preço_origem._style)
            data_destino._style = copy(data_origem._style)
            nome_destino._style = copy(nome_origem._style)
            arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
        else:
            price = price - (price * 0.1)
            totdesc += 1
            total += price
            totprod += 1
            sacola.append((name, price))
            banco.append([name,price,data])
            preço_destino = banco.cell(row=ultima_linha+1, column=2)
            data_destino = banco.cell(row=ultima_linha+1, column=3)
            nome_destino = banco.cell(row=ultima_linha+1, column=1)
            preço_destino._style = copy(preço_origem._style)
            data_destino._style = copy(data_origem._style)
            nome_destino._style = copy(nome_origem._style)
            arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
        while True:
            print(f"{"\n"}{"\033[1;36m ADD MORE PRODUCTS? ".center(48)}{"\n"}")
            decide = input(f"{" [1]YES   OR   [2]NO ".center(42)}{"\n"}{"\n"}").strip().lower()
            if decide == "2":
                print(f"{"\n"}{"\033[1;33m SUMMARY OF YOUR SALE ".center(48)}")
                print(
                    f"{"\n"}{"\033[1;33mTotal products: ".ljust(35)}\033[1;32m{totprod:.1f}\033[m")
                print(
                    f"{"\033[1;33mTotal products on sale: ".ljust(35)}\033[1;32m{totdesc:.1f}\033[m")
                print(
                    f"{"\033[1;33mTotal to pay: ".ljust(35)}\033[1;32mR${total:.2f}\033[m")
                print(f"{"\n"}{"\033[1;33m ALL PRODUCTS SOLD ".center(47)}{"\n"}")
                for name, price in sacola:
                    print(f"\033[1;33m{name}{"    -    "}\033[1;32mR${price:.2f}")
                print(f"{"\n"}{"\033[m"}")
                arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                break
            elif decide == "1":
                arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
                break
            else:
                print(
                    f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT ".center(50)}")
                print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(52)}{"\n"}")
        if decide != "1" and "2":
            break
    


# vender_prod()
