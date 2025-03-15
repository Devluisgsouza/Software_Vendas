import openpyxl as op
from copy import copy

exit = False

def criar_conta():
    arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
    banco_contas = arquivo_bancodd["contas_de_usuario"]
    ultima_linha = banco_contas.max_row
    for linha in range(2, ultima_linha +1):
        ultima_linha = linha
    id_origem = banco_contas.cell(row=2, column=1)
    email_origem = banco_contas.cell(row=2, column=2)
    senha_origem = banco_contas.cell(row=2, column=3)
    print(f"{"\n"}{"\033[1;36m CREATE YOUR ACCOUNT ".center(52)}{"\n"}")
    while True:
        if exit:
            break
        login = input("\033[1;32mEmail: ").strip().lower()
        password = input("\033[1;32mPassword: ").strip()
        for linha in range(2,ultima_linha + 1):
            emails = banco_contas.cell(row=linha, column=2).value
            id = banco_contas.cell(row=linha, column=1).value
            id_conta = int(id)  + 1
            if login == emails:
                print(f"{"\n"}{"\033[1;31m THIS ACCOUNT ALREADY EXISTS IN THE SYSTEM \033[m".center(52, "=")}{"\n"}")
                break
        else:
            print(f"{"\n"}{"\033[1;32m ACCOUNT CREATED SUCCESSFULY! \033[m".center(54)}")
            banco_contas.append([id_conta,login,password])
            email_destino = banco_contas.cell(row=ultima_linha+1, column=2)
            senha_destino = banco_contas.cell(row=ultima_linha+1, column=3)
            id_destino = banco_contas.cell(row=ultima_linha+1, column=1)
            email_destino._style = copy(email_origem._style)
            senha_destino._style = copy(senha_origem._style)
            id_destino._style = copy(id_origem._style)
            arquivo_bancodd.save("main\\banco_de_dados\\banco.xlsx")
            return exit
 

def fazer_login():
    arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
    banco_contas = arquivo_bancodd["contas_de_usuario"]
    ultima_linha = banco_contas.max_row
    print(f"{"\n"}{"\033[1;36m ENTER WITH YOUR ACCOUNT: ".center(52)}{"\n"}")
    login = input("\033[1;32mEmail: ").strip().lower()
    password = input("Password: ").strip()
    for linha in range(2,ultima_linha + 1):
        emails = banco_contas.cell(row=linha, column=1).value
        if login == emails:
            numero_linha_logins = banco_contas.cell(row=linha, column=2).row
            senhas = banco_contas.cell(row=numero_linha_logins, column=3).value
            if password == senhas:
                print(f"{"\n"}{"\033[1;32m SUCCESSFULLY LOGIN \033[m".center(51)}")
                return
    print(f"{"\n"}{"\033[1;31m WRONG INFORMATIONS! \033[m".center(55)}")
    print(f"\033[1;31mTHIS ACCOUNT DOESN'T EXIST IN THE SYSTEM".center(52))
    while True:
        print(f"{"\n"}{"\033[1;36m WOULD YOU LIKE TO CREATE A NEW ACCOUNT? ".center(53)}{"\n"}")
        resp = input(f"{" [1]YES   OR   [2]NO ".center(44)}{"\n"}{"\n"}").replace(" ","").lower()
        if resp == "1":
            criar_conta()
            break
        elif resp == "2":
            fazer_login()
            break
        else:
            print(f"{"\n"}{"\033[1;31mI CAN'T UNDERSTAND WHAT YOU WANT\033[m".center(53)}")
            print(f"{"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2]".center(53)}{"\n"}")
            

# criar_conta()
# fazer_login()